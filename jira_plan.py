#!/usr/bin/env python3
"""
Normalize Jira CSV/XLSX exports into:
1. A workbook with tracking + dev-plan sheets
2. A machine-friendly JSON summary for agent orchestration

This script stays intentionally repo-agnostic so it can be copied into any
AEM codebase as part of `.agent/universal/`.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from copy import deepcopy
from pathlib import Path
from typing import Iterable, Iterator

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - handled at runtime for CSV-only use
    openpyxl = None
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


TRACKING_HEADERS = [
    "Agent - Summary",
    "Agent - Status",
    "Agent - Notes",
    "Agent - Branch",
    "Agent - PR URL",
]

PLAN_HEADERS = [
    "Jira Key",
    "Title",
    "Type",
    "Priority",
    "Acceptance Criteria",
    "Assumptions / Missing Info",
    "Repo Areas Impacted",
    "Existing Code References",
    "Implementation Steps",
    "Test Plan",
    "Rollout / Feature Flag Notes",
    "Risk & Edge Cases",
    "Estimated Complexity",
    "Dependencies / Ordering",
    "Prompt for aem-feature",
    "Prompt for Copilot/Claude",
    "Prompt Notes / Guardrails",
    "Planned Branch Name",
    "Planned PR Title",
    "Planned PR Body",
]

CANONICAL_COLUMNS = {
    "jira_key": {"jira key", "issue key", "key", "id"},
    "summary": {"summary", "title", "issue summary"},
    "description": {"description", "details"},
    "acceptance_criteria": {
        "acceptance criteria",
        "acceptance criterion",
        "acceptance",
        "ac",
        "criteria",
    },
    "issue_type": {"issue type", "type"},
    "priority": {"priority"},
    "labels": {"labels", "label", "tags"},
    "components": {"components", "component", "component/s"},
    "epic_link": {"epic link", "epic", "epic key", "parent"},
    "story_points": {
        "story points",
        "story point",
        "story point estimate",
        "points",
    },
}

DEFAULT_ROW = {key: "" for key in CANONICAL_COLUMNS}
TEXT_WRAP = None
THIN_BORDER = None
HEADER_FONT = None
HEADER_FILL = None


def init_styles() -> None:
    global TEXT_WRAP, THIN_BORDER, HEADER_FONT, HEADER_FILL
    if openpyxl is None:
        return
    TEXT_WRAP = Alignment(wrap_text=True, vertical="top")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )


def slugify(value: str, fallback: str = "story") -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or fallback


def normalize_header(value: str) -> str:
    value = value or ""
    value = value.strip().lower()
    value = re.sub(r"[_/\\-]+", " ", value)
    value = re.sub(r"[^a-z0-9 ]+", "", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def find_header_map(headers: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for canonical, aliases in CANONICAL_COLUMNS.items():
            if normalized in aliases and canonical not in header_map:
                header_map[canonical] = index
                break
    return header_map


def find_header_row(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    best_index = -1
    best_map: dict[str, int] = {}
    best_score = -1
    for row_index, row in enumerate(rows[:25]):
        row_strings = [str(cell).strip() for cell in row]
        header_map = find_header_map(row_strings)
        score = len(header_map)
        if "jira_key" in header_map and "summary" in header_map:
            score += 3
        if score > best_score:
            best_index = row_index
            best_map = header_map
            best_score = score
    if best_index < 0 or "jira_key" not in best_map or "summary" not in best_map:
        raise ValueError(
            "Could not find a Jira header row with at least Jira Key and Summary columns."
        )
    return best_index, best_map


def as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def iter_csv_rows(path: Path) -> Iterator[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            yield [as_text(cell) for cell in row]


def load_xlsx_rows(path: Path) -> tuple[str, list[list[str]]]:
    if openpyxl is None:
        raise RuntimeError(
            "XLSX support requires openpyxl. Install it with: pip install openpyxl"
        )
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    best_sheet = None
    best_rows: list[list[str]] = []
    best_score = -1

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = [[as_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
        try:
            _, header_map = find_header_row(rows)
        except ValueError:
            continue
        score = len(header_map)
        if score > best_score:
            best_score = score
            best_sheet = sheet_name
            best_rows = rows

    if best_sheet is None:
        raise ValueError("Could not find a Jira-like sheet in the workbook.")
    return best_sheet, best_rows


def load_rows(path: Path) -> tuple[str, list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "CSV", list(iter_csv_rows(path))
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_rows(path)
    if suffix == ".xls":
        raise RuntimeError(
            "Legacy .xls files are not supported directly. Export to .xlsx or .csv first."
        )
    raise RuntimeError(f"Unsupported Jira export format: {path.suffix}")


def normalize_row(
    row: list[str],
    header_map: dict[str, int],
    source_sheet: str,
    source_row_number: int,
    branch_prefix: str,
) -> dict[str, object]:
    normalized = deepcopy(DEFAULT_ROW)
    for canonical, index in header_map.items():
        normalized[canonical] = row[index].strip() if index < len(row) else ""

    summary = normalized["summary"]
    description = normalized["description"]
    acceptance_criteria = normalized["acceptance_criteria"]
    jira_key = normalized["jira_key"] or f"ROW-{source_row_number}"
    story_slug = slugify(summary or jira_key, fallback=jira_key.lower())
    prefix = branch_prefix.strip().strip("/") or "feature"
    branch_slug = slugify(f"{jira_key}-{summary}", fallback=jira_key.lower())

    normalized.update(
        {
            "source_sheet": source_sheet,
            "source_row": source_row_number,
            "story_slug": story_slug,
            "agent_summary": build_agent_summary(summary, description, acceptance_criteria),
            "agent_status": "Planned",
            "agent_notes": "Needs repo grounding and validation.",
            "agent_branch": f"{prefix}/{branch_slug}",
            "agent_pr_url": "",
            "planned_pr_title": f"[{jira_key}] {summary}".strip(),
        }
    )
    return normalized


def build_agent_summary(summary: str, description: str, acceptance_criteria: str) -> str:
    seed = summary or description or acceptance_criteria or "Jira story"
    seed = re.sub(r"\s+", " ", seed).strip()
    if len(seed) <= 160:
        return seed
    return seed[:157].rstrip() + "..."


def normalize_export(path: Path, branch_prefix: str) -> dict[str, object]:
    source_sheet, rows = load_rows(path)
    if not rows:
        raise ValueError("The Jira export is empty.")

    header_row_index, header_map = find_header_row(rows)
    stories: list[dict[str, object]] = []
    duplicates: list[str] = []
    seen_keys: set[str] = set()

    for offset, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        if not any(cell.strip() for cell in row):
            continue
        normalized = normalize_row(row, header_map, source_sheet, offset, branch_prefix)
        jira_key = str(normalized["jira_key"]).strip()
        if not jira_key:
            continue
        if jira_key in seen_keys:
            duplicates.append(jira_key)
            continue
        seen_keys.add(jira_key)
        stories.append(normalized)

    missing_ac = [story["jira_key"] for story in stories if not story["acceptance_criteria"]]
    issue_type_counts = Counter(
        story["issue_type"] or "Unspecified" for story in stories  # type: ignore[arg-type]
    )

    return {
        "source_file": str(path),
        "source_sheet": source_sheet,
        "header_row": header_row_index + 1,
        "header_map": header_map,
        "total_items": len(stories),
        "duplicates_skipped": duplicates,
        "missing_acceptance_criteria": missing_ac,
        "issue_type_counts": dict(sorted(issue_type_counts.items())),
        "stories": stories,
    }


def apply_filters(
    stories: Iterable[dict[str, object]],
    priorities: set[str],
    issue_types: set[str],
    include_missing_ac: bool,
) -> list[dict[str, object]]:
    filtered = []
    for story in stories:
        priority = str(story["priority"]).strip().lower()
        issue_type = str(story["issue_type"]).strip().lower()
        has_ac = bool(str(story["acceptance_criteria"]).strip())

        if priorities and priority not in priorities:
            continue
        if issue_types and issue_type not in issue_types:
            continue
        if not include_missing_ac and not has_ac:
            story = dict(story)
            story["agent_status"] = "Needs Clarification"
            story["agent_notes"] = "Missing acceptance criteria in source export."
        filtered.append(story)
    return filtered


def count_issue_types(stories: Iterable[dict[str, object]]) -> dict[str, int]:
    counts = Counter(str(story["issue_type"]).strip() or "Unspecified" for story in stories)
    return dict(sorted(counts.items()))


def ensure_parent(path: Path) -> None:
    if path.parent and not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)


def write_json(path: Path, payload: dict[str, object]) -> None:
    ensure_parent(path)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def set_cell(worksheet, row: int, col: int, value: object) -> None:
    cell = worksheet.cell(row=row, column=col, value=value)
    if TEXT_WRAP is not None:
        cell.alignment = TEXT_WRAP
    if THIN_BORDER is not None:
        cell.border = THIN_BORDER


def style_header(cell) -> None:
    if HEADER_FONT is not None:
        cell.font = HEADER_FONT
    if HEADER_FILL is not None:
        cell.fill = HEADER_FILL
    if TEXT_WRAP is not None:
        cell.alignment = TEXT_WRAP
    if THIN_BORDER is not None:
        cell.border = THIN_BORDER


def build_aem_feature_prompt(story: dict[str, object]) -> str:
    key = story["jira_key"]
    summary = story["summary"]
    description = story["description"] or "No description provided."
    acceptance = story["acceptance_criteria"] or "No acceptance criteria provided."
    branch = story["agent_branch"]
    return f"""Read `.agent/project.yaml`, `.agent/AGENT.md`, and `.agent/REPO_CONTEXT.md` fully.
Then implement this Jira story end to end on branch `{branch}`.

Jira Key: {key}
Summary: {summary}
Description:
{description}

Acceptance Criteria:
{acceptance}

Required output:
- Files created/updated
- Build/test results with real command output summaries
- Test page URL and authoring URL
- Any blockers or assumptions
- If the test-page root is unclear, ask for the desired language/content root and create `/test-pages` beneath it
"""


def build_copilot_prompt(story: dict[str, object]) -> str:
    key = story["jira_key"]
    summary = story["summary"]
    acceptance = story["acceptance_criteria"] or "Clarify missing acceptance criteria."
    return f"""Implement Jira story {key}: {summary}

Use the repo patterns already present in the AEM codebase.
Acceptance criteria:
{acceptance}

Add tests, avoid unrelated refactors, and summarize changed files plus validation results.
"""


def build_pr_body(story: dict[str, object]) -> str:
    summary = story["summary"]
    acceptance = str(story["acceptance_criteria"]).strip() or "Acceptance criteria not provided."
    return f"""## Summary
{summary}

## Acceptance Criteria
{acceptance}

## Validation
- Build: pending
- Tests: pending
"""


def build_workbook(path: Path, payload: dict[str, object]) -> None:
    if openpyxl is None or Workbook is None:
        raise RuntimeError(
            "Workbook generation requires openpyxl. Install it with: pip install openpyxl"
        )

    init_styles()
    workbook = Workbook()
    import_sheet = workbook.active
    import_sheet.title = "Jira Import"

    import_headers = [
        "Jira Key",
        "Issue Type",
        "Priority",
        "Summary",
        "Description",
        "Acceptance Criteria",
        "Labels",
        "Components",
        "Epic Link",
        "Story Points",
        "Source Sheet",
        "Source Row",
    ] + TRACKING_HEADERS

    for index, header in enumerate(import_headers, start=1):
        cell = import_sheet.cell(row=1, column=index, value=header)
        style_header(cell)

    stories = payload["stories"]
    for row_index, story in enumerate(stories, start=2):
        values = [
            story["jira_key"],
            story["issue_type"],
            story["priority"],
            story["summary"],
            story["description"],
            story["acceptance_criteria"],
            story["labels"],
            story["components"],
            story["epic_link"],
            story["story_points"],
            story["source_sheet"],
            story["source_row"],
            story["agent_summary"],
            story["agent_status"],
            story["agent_notes"],
            story["agent_branch"],
            story["agent_pr_url"],
        ]
        for col_index, value in enumerate(values, start=1):
            set_cell(import_sheet, row_index, col_index, value)

    dev_plan_sheet = workbook.create_sheet("Dev Plan (Agent Output)")
    for index, header in enumerate(PLAN_HEADERS, start=1):
        cell = dev_plan_sheet.cell(row=1, column=index, value=header)
        style_header(cell)

    for row_index, story in enumerate(stories, start=2):
        values = [
            story["jira_key"],
            story["summary"],
            story["issue_type"],
            story["priority"],
            story["acceptance_criteria"],
            "Fill after repo analysis.",
            "Fill with concrete file paths and symbols.",
            "Fill with repo patterns/components to follow.",
            "Break into repo-specific implementation steps.",
            "Add module-level build/test commands and any visual QA steps.",
            "Document deployment profile, content package, or feature flag needs.",
            "Capture authoring edge cases, dispatcher concerns, and content assumptions.",
            "",
            "",
            build_aem_feature_prompt(story),
            build_copilot_prompt(story),
            "Do not refactor unrelated areas. Reuse repo patterns. Add or update policies only when required.",
            story["agent_branch"],
            story["planned_pr_title"],
            build_pr_body(story),
        ]
        for col_index, value in enumerate(values, start=1):
            set_cell(dev_plan_sheet, row_index, col_index, value)

    width_overrides = {
        "A": 16,
        "B": 18,
        "C": 12,
        "D": 40,
        "E": 50,
        "F": 50,
        "G": 22,
        "H": 22,
        "I": 18,
        "J": 12,
        "K": 16,
        "L": 12,
        "M": 40,
        "N": 18,
        "O": 24,
        "P": 30,
        "Q": 30,
        "R": 30,
        "S": 30,
        "T": 40,
    }
    for worksheet in (import_sheet, dev_plan_sheet):
        for column, width in width_overrides.items():
            worksheet.column_dimensions[column].width = width
        worksheet.freeze_panes = "A2"

    ensure_parent(path)
    workbook.save(path)


def parse_list_argument(value: str | None) -> set[str]:
    if not value:
        return set()
    items = [item.strip().lower() for item in value.split(",")]
    return {item for item in items if item}


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_with_dev_plan.xlsx")


def default_json_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_normalized.json")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Normalize Jira CSV/XLSX exports for AEM delivery agents."
    )
    parser.add_argument("input", help="Path to the Jira CSV/XLSX export")
    parser.add_argument(
        "--output",
        help="Path to the generated XLSX workbook (defaults beside the input file)",
    )
    parser.add_argument(
        "--json",
        dest="json_output",
        help="Path to the generated JSON summary (defaults beside the input file)",
    )
    parser.add_argument(
        "--priority",
        help="Comma-separated priority filter, e.g. High,Highest,2 - High",
    )
    parser.add_argument(
        "--issue-type",
        help="Comma-separated issue type filter, e.g. Story,Bug,Task",
    )
    parser.add_argument(
        "--branch-prefix",
        default="feature",
        help="Branch prefix for generated per-story branches (default: feature)",
    )
    parser.add_argument(
        "--include-missing-ac",
        action="store_true",
        help="Keep missing acceptance criteria as Planned instead of Needs Clarification",
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        parser.error(f"Input file does not exist: {input_path}")

    try:
        payload = normalize_export(input_path, args.branch_prefix)
        filtered_stories = apply_filters(
            payload["stories"],
            parse_list_argument(args.priority),
            parse_list_argument(args.issue_type),
            args.include_missing_ac,
        )
        payload["stories"] = filtered_stories
        payload["total_items"] = len(filtered_stories)
        payload["missing_acceptance_criteria"] = [
            story["jira_key"] for story in filtered_stories if not story["acceptance_criteria"]
        ]
        payload["issue_type_counts"] = count_issue_types(filtered_stories)

        workbook_path = Path(args.output).expanduser() if args.output else default_output_path(input_path)
        json_path = Path(args.json_output).expanduser() if args.json_output else default_json_path(input_path)
        workbook_path = workbook_path.resolve()
        json_path = json_path.resolve()

        write_json(json_path, payload)
        build_workbook(workbook_path, payload)

        summary = {
            "workbook": str(workbook_path),
            "json": str(json_path),
            "stories": payload["total_items"],
            "missing_acceptance_criteria": len(payload["missing_acceptance_criteria"]),
            "duplicates_skipped": len(payload["duplicates_skipped"]),
            "issue_type_counts": payload["issue_type_counts"],
        }
        print(json.dumps(summary, indent=2))
        return 0
    except Exception as exc:  # pragma: no cover - runtime UX path
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
